import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import random

# Create workbook for Hotel
wb_hotel = openpyxl.Workbook()
ws_hotel = wb_hotel.active
ws_hotel.title = "Hotel Service Fee"

# Header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Hotel Headers
hotel_headers = [
    "No",
    "Transaction Time",
    "Booking ID",
    "Status",
    "Hotel Name",
    "Room Type",
    "Employee Name",
    "Transaction Amount (Rp)",
    "Service Fee (Rp)",
    "Sheet",
    "Settlement Method",
    "Currency"
]

for col, header in enumerate(hotel_headers, 1):
    cell = ws_hotel.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample hotel data
hotels = [
    ("Aston Hotel Makassar", "Deluxe Twin"),
    ("Swiss-Belhotel Makassar", "Superior Queen"),
    ("MaxOne Hotels Makassar", "Standard Double"),
    ("The Rinra Makassar", "Executive Suite"),
    ("Harper Hotel Makassar", "Premier King")
]

employees = [
    "BUDI SANTOSO",
    "DEWI LESTARI",
    "AHMAD FADLI",
    "SRI RAHAYU",
    "RUDI HARTONO"
]

for i in range(5):
    hotel_name, room_type = hotels[i]
    booking_id = f"99{random.randint(10000000, 99999999)}"
    transaction_time = f"15 Dec 2025, {random.randint(10, 23)}:{random.randint(10, 59)}:00"
    transaction_amount = random.randint(500000, 1500000)
    service_fee = int(transaction_amount * 0.01)
    
    row = [
        i + 1,
        transaction_time,
        booking_id,
        "ISSUED",
        hotel_name,
        room_type,
        employees[i],
        transaction_amount,
        service_fee,
        "Desember 2025",
        "INVOICE",
        "IDR"
    ]
    
    for col, value in enumerate(row, 1):
        ws_hotel.cell(row=i+2, column=col, value=value)

# Adjust column widths
ws_hotel.column_dimensions['A'].width = 8
ws_hotel.column_dimensions['B'].width = 22
ws_hotel.column_dimensions['C'].width = 15
ws_hotel.column_dimensions['D'].width = 10
ws_hotel.column_dimensions['E'].width = 35
ws_hotel.column_dimensions['F'].width = 20
ws_hotel.column_dimensions['G'].width = 25
ws_hotel.column_dimensions['H'].width = 20
ws_hotel.column_dimensions['I'].width = 18
ws_hotel.column_dimensions['J'].width = 18
ws_hotel.column_dimensions['K'].width = 18
ws_hotel.column_dimensions['L'].width = 10

wb_hotel.save("data/sample_service_fee_hotel.xlsx")
print("✓ Created data/sample_service_fee_hotel.xlsx with 5 hotel bookings")

# Create workbook for Flight
wb_flight = openpyxl.Workbook()
ws_flight = wb_flight.active
ws_flight.title = "Flight Service Fee"

# Flight Headers
flight_headers = [
    "No",
    "Transaction Time",
    "Booking ID",
    "Status",
    "Route",
    "Trip Type",
    "Pax",
    "Airline ID",
    "Booker Email",
    "Passenger Name (Employee)",
    "Transaction Amount (Rp)",
    "Service Fee (Rp)",
    "Sheet",
    "Settlement Method",
    "Currency"
]

for col, header in enumerate(flight_headers, 1):
    cell = ws_flight.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample flight data
airlines = ["GA", "JT", "ID", "QG", "IU"]

for i in range(5):
    booking_id = f"99{random.randint(10000000, 99999999)}"
    transaction_time = f"15 Dec 2025, {random.randint(6, 23)}:{random.randint(10, 59)}:00"
    transaction_amount = random.randint(800000, 2500000)
    service_fee = int(transaction_amount * 0.01)
    
    row = [
        i + 1,
        transaction_time,
        booking_id,
        "ISSUED",
        "ONE_WAY",
        "ONE_WAY",
        1,
        airlines[i],
        f"{employees[i].lower().replace(' ', '.')}@pln.co.id",
        employees[i],
        transaction_amount,
        service_fee,
        "Desember 2025",
        "INVOICE",
        "IDR"
    ]
    
    for col, value in enumerate(row, 1):
        ws_flight.cell(row=i+2, column=col, value=value)

# Adjust column widths
ws_flight.column_dimensions['A'].width = 8
ws_flight.column_dimensions['B'].width = 22
ws_flight.column_dimensions['C'].width = 15
ws_flight.column_dimensions['D'].width = 10
ws_flight.column_dimensions['E'].width = 12
ws_flight.column_dimensions['F'].width = 12
ws_flight.column_dimensions['G'].width = 8
ws_flight.column_dimensions['H'].width = 12
ws_flight.column_dimensions['I'].width = 30
ws_flight.column_dimensions['J'].width = 25
ws_flight.column_dimensions['K'].width = 20
ws_flight.column_dimensions['L'].width = 18
ws_flight.column_dimensions['M'].width = 18
ws_flight.column_dimensions['N'].width = 18
ws_flight.column_dimensions['O'].width = 10

wb_flight.save("data/sample_service_fee_flight.xlsx")
print("✓ Created data/sample_service_fee_flight.xlsx with 5 flight bookings")

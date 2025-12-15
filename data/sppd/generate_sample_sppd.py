import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def generate_sample_sppd_excel():
    """
    Generate sample SPPD Excel file with planned payment date column
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "November 2025"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 15,  # Trip Number
        'B': 12,  # Group ID
        'C': 12,  # Additional
        'D': 25,  # Customer Name
        'E': 12,  # Additional
        'F': 30,  # Trip Destination
        'G': 40,  # Reason for Trip
        'H': 15,  # Trip Begins On
        'I': 15,  # Trip Ends On
        'J': 18,  # Planned Payment Date (TANGGAL BAYAR)
        'K': 12,  # Company Code
        'L': 15,  # Paid Amount
        'M': 10,  # Currency
        'N': 12,  # Additional
        'O': 25,  # Beneficiary Bank Name
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Headers (Row 1)
    headers = [
        'Trip Number',
        'Group ID',
        'Cost Center',
        'Customer Name',
        'Position',
        'Trip Destination',
        'Reason for Trip',
        'Trip Begins On',
        'Trip Ends On',
        'Tanggal Bayar',  # KOLOM BARU!
        'Company Code',
        'Paid Amount',
        'Currency',
        'Account Number',
        'Beneficiary Bank Name'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Sample data
    sample_names = [
        'BUDIMAN',
        'RIFKI MAULANA MUSLIM',
        'IBNU RUSYDI',
        'MUH. IQRA WAHYUDI',
        'ALFIAN ROMADHAN',
        'SITI NURHALIZA',
        'AHMAD FAUZI',
        'DEWI LESTARI',
        'BAMBANG SURYANTO',
        'NURUL HIDAYAH'
    ]
    
    sample_destinations = [
        'Kota Parepare - Kab. Soppeng',
        'Kab. Bulukumba - Kota Makassar',
        'Kota Makassar - Kab. Gowa',
        'Kab. Sinjai - Kota Parepare',
        'Kota Palopo - Kab. Luwu',
        'Kab. Bone - Kota Makassar',
        'Kab. Wajo - Kab. Soppeng',
        'Kota Makassar - Kab. Takalar'
    ]
    
    sample_reasons = [
        'Mutasi Pegawai',
        'Mutasi Jabatan',
        'Orientasi Mutasi Jabatan',
        'Penugasan Pelaksanaan Syukuran Wisuda Periode 24-1',
        'Menghadiri Rapat di UP2D SULSELRABAR',
        'Keandalan dan efisiensi distribusi',
        'Koordinasi Proyek PLN',
        'Audit Internal Unit'
    ]
    
    sample_banks = [
        'Bank Mandiri',
        'Bank BNI',
        'Bank BRI',
        'Bank BTN',
        'Bank Syariah Indonesia',
        'Bank CIMB Niaga'
    ]
    
    # Generate 15 sample rows
    start_date = datetime(2025, 11, 1)
    row_num = 2
    
    for i in range(15):
        trip_number = f'412017{7500 + i}'
        group_id = f'G{1000 + i}'
        cost_center = f'CC{2000 + i}'
        customer_name = random.choice(sample_names)
        position = random.choice(['Staff', 'Supervisor', 'Manager', 'Asisten Manager'])
        destination = random.choice(sample_destinations)
        reason = random.choice(sample_reasons)
        
        # Random trip dates in November 2025
        trip_start = start_date + timedelta(days=random.randint(0, 28))
        trip_duration = random.randint(1, 5)
        trip_end = trip_start + timedelta(days=trip_duration)
        
        # Planned payment date: biasanya 3-10 hari setelah trip selesai
        payment_date = trip_end + timedelta(days=random.randint(3, 10))
        
        company_code = 'PLN001'
        paid_amount = random.randint(500000, 5000000)
        currency = 'IDR'
        account_number = f'{random.randint(1000000000, 9999999999)}'
        bank_name = random.choice(sample_banks)
        
        # Write data
        row_data = [
            trip_number,
            group_id,
            cost_center,
            customer_name,
            position,
            destination,
            reason,
            trip_start,  # Will be formatted as date
            trip_end,    # Will be formatted as date
            payment_date,  # KOLOM TANGGAL BAYAR!
            company_code,
            paid_amount,
            currency,
            account_number,
            bank_name
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = border
            
            # Format date columns
            if col_idx in [8, 9, 10]:  # Trip Begins On, Trip Ends On, Tanggal Bayar
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = Alignment(horizontal="center")
            
            # Format amount column
            if col_idx == 12:  # Paid Amount
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            
            # Center align for some columns
            if col_idx in [1, 2, 3, 5, 10, 11, 13]:
                cell.alignment = Alignment(horizontal="center")
        
        row_num += 1
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    output_file = r"d:\Bu Intan\Bu Intan\data\sppd\sample_sppd_with_payment_date.xlsx"
    wb.save(output_file)
    print(f"✓ Sample SPPD Excel created: {output_file}")
    print(f"✓ Total rows: {row_num - 1} (excluding header)")
    print(f"\nKolom yang tersedia:")
    print("  - Column A: Trip Number")
    print("  - Column D: Customer Name")
    print("  - Column F: Trip Destination")
    print("  - Column G: Reason for Trip")
    print("  - Column H: Trip Begins On (Tanggal Mulai)")
    print("  - Column I: Trip Ends On (Tanggal Selesai)")
    print("  - Column J: Tanggal Bayar (PLANNED PAYMENT DATE) ← BARU!")
    print("  - Column L: Paid Amount")
    print("  - Column O: Beneficiary Bank Name")
    print(f"\n📋 File siap untuk di-convert dengan convert_sppd_excel.py")

if __name__ == "__main__":
    generate_sample_sppd_excel()

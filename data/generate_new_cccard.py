import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import random

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CC Card Transactions"

# Header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Headers
headers = [
    "No.",
    "Booking ID", 
    "Name",
    "Personel Number",
    "Trip Number",
    "Trip Destination",
    "Trip Date",
    "Payment",
    "Transaction Type"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample data - NEW booking IDs that don't exist yet
names = [
    "BUDI SANTOSO",
    "SRI RAHAYU", 
    "AHMAD FAUZI",
    "DEWI LESTARI",
    "RUDI HARTONO"
]

cities = [
    ["Kota Jakarta", "Kota Surabaya"],
    ["Kota Bandung", "Kota Medan"],
    ["Kota Malang", "Kota Denpasar"],
    ["Kota Semarang", "Kota Padang"],
    ["Kota Yogyakarta", "Kota Palembang"]
]

# Generate new transactions
for i in range(5):
    # Generate unique booking ID (9999xxxxxx - high number unlikely to exist)
    booking_id = f"9999{random.randint(100000, 999999)}"
    personel_number = f"{random.randint(90000000, 99999999)}"
    trip_number = f"4120{random.randint(100000, 999999)}"
    
    origin, destination = cities[i]
    trip_dest = f"{origin} - {destination}"
    
    # Random dates in December 2025
    dep_day = random.randint(1, 25)
    dep_date = datetime(2025, 12, dep_day)
    duration = random.randint(2, 5)
    ret_date = dep_date + timedelta(days=duration)
    
    trip_date = f"{dep_date.strftime('%d/%m/%Y')} - {ret_date.strftime('%d/%m/%Y')}"
    
    # Random payment amount
    payment = random.randint(300000, 1000000)
    
    row = [
        i + 1,
        booking_id,
        names[i],
        personel_number,
        trip_number,
        trip_dest,
        trip_date,
        payment,
        "payment"
    ]
    
    for col, value in enumerate(row, 1):
        ws.cell(row=i+2, column=col, value=value)

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 15

# Save
wb.save("data/new_cccard_test.xlsx")
print("✓ Created data/new_cccard_test.xlsx with 5 new transactions")
print("  Booking IDs start with 9999 to ensure they are new")

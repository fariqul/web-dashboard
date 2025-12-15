import openpyxl
from datetime import datetime
import csv
import re

def clean_trip_number(trip_num):
    """
    Clean and preserve original trip number from Excel
    """
    if not trip_num:
        return ""
    # Convert to string and strip whitespace
    trip_num = str(trip_num).strip()
    # Remove any decimal points (e.g., "4120178765.0" -> "4120178765")
    if '.' in trip_num:
        trip_num = trip_num.split('.')[0]
    return trip_num

def clean_amount(amount):
    """
    Clean currency amount
    """
    if not amount:
        return "0"
    # Convert to string
    amount_str = str(amount).strip()
    # Remove currency symbols, commas, and dots
    amount_str = re.sub(r'[Rp,.\s]', '', amount_str)
    # If empty after cleaning, return 0
    return amount_str if amount_str else "0"

def get_cell_value(cell):
    """
    Get cell value and handle dates properly
    """
    if cell.value is None:
        return None
    # Check if cell is formatted as date
    if isinstance(cell.value, datetime):
        return cell.value
    elif cell.number_format and ('d' in cell.number_format.lower() or 'm' in cell.number_format.lower() or 'y' in cell.number_format.lower()):
        # Cell has date formatting, try to convert
        if isinstance(cell.value, (int, float)):
            return openpyxl.utils.datetime.from_excel(cell.value)
    return cell.value

def excel_date_to_string(date_value):
    """
    Convert date value to YYYY-MM-DD string
    """
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    elif isinstance(date_value, (int, float)) and date_value > 0:
        # Try to convert as Excel date
        try:
            excel_date = openpyxl.utils.datetime.from_excel(date_value)
            return excel_date.strftime('%Y-%m-%d')
        except:
            pass
    if date_value and str(date_value).strip():
        return str(date_value).strip()
    return ""

def convert_excel_to_sppd_csv(excel_path, output_csv):
    """
    Convert Excel SPPD data to standardized CSV format
    Preserves original trip numbers from Excel
    """
    # Load the workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # CSV header
    header = [
        'trip_number', 'customer_name', 'trip_destination', 'reason_for_trip',
        'trip_begins_on', 'trip_ends_on', 'planned_payment_date', 'paid_amount', 'beneficiary_bank_name'
    ]
    
    all_rows = []
    
    # Process Sheet1 ONLY (not Sheet1 (2))
    sheets_to_process = []
    if 'Sheet1' in wb.sheetnames:
        sheets_to_process = ['Sheet1']
    else:
        sheets_to_process = wb.sheetnames
    
    for sheet_name in sheets_to_process:
        sheet = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Find header row by looking for "Trip Number"
        header_row = None
        for idx in range(1, 11):
            row_cells = list(sheet[idx])
            row_values = [str(cell.value) if cell.value else "" for cell in row_cells]
            if 'Trip Number' in row_values:
                header_row = idx
                print(f"Found header at row {header_row}")
                break
        
        if not header_row:
            print(f"Skipping sheet {sheet_name} - no header found")
            continue
        
        # Get column indices from Excel header row
        excel_header = [str(cell.value).strip() if cell.value else "" for cell in sheet[header_row]]
        print(f"Excel header found: {excel_header[:10]}")  # Debug: show first 10 columns
        
        # Map columns from Excel header
        col_trip_number = excel_header.index('Trip Number') if 'Trip Number' in excel_header else None
        col_customer_name = excel_header.index('Customer Name') if 'Customer Name' in excel_header else None
        col_trip_destination = excel_header.index('Trip Destination') if 'Trip Destination' in excel_header else None
        col_reason = excel_header.index('Reason for Trip') if 'Reason for Trip' in excel_header else None
        col_begins = excel_header.index('Trip Begins On') if 'Trip Begins On' in excel_header else None
        col_ends = excel_header.index('Trip Ends On') if 'Trip Ends On' in excel_header else None
        col_planned_payment = excel_header.index('Tanggal Rencana Bayar') if 'Tanggal Rencana Bayar' in excel_header else None
        if col_planned_payment is None:
            col_planned_payment = excel_header.index('Tanggal Bayar') if 'Tanggal Bayar' in excel_header else None
        col_paid_amount = excel_header.index('Paid Amount') if 'Paid Amount' in excel_header else None
        col_bank = excel_header.index('Beneficiary Bank Name') if 'Beneficiary Bank Name' in excel_header else None
        
        print(f"Column mapping: trip_number={col_trip_number}, customer={col_customer_name}, destination={col_trip_destination}")
        
        # Process data rows
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_cells = list(sheet[row_idx])
            
            # Skip empty rows
            if not any(cell.value for cell in row_cells):
                continue
                
            try:
                # Extract data using column indices with proper cell access
                trip_number = clean_trip_number(get_cell_value(row_cells[col_trip_number])) if col_trip_number is not None else ""
                
                # Skip if no trip number
                if not trip_number:
                    continue
                
                customer_name = str(get_cell_value(row_cells[col_customer_name])).strip() if col_customer_name is not None and get_cell_value(row_cells[col_customer_name]) else ""
                trip_destination = str(get_cell_value(row_cells[col_trip_destination])).strip() if col_trip_destination is not None and get_cell_value(row_cells[col_trip_destination]) else ""
                reason_for_trip = str(get_cell_value(row_cells[col_reason])).strip() if col_reason is not None and get_cell_value(row_cells[col_reason]) else ""
                
                # Convert dates using excel_date_to_string function
                trip_begins_on = excel_date_to_string(get_cell_value(row_cells[col_begins])) if col_begins is not None else ""
                trip_ends_on = excel_date_to_string(get_cell_value(row_cells[col_ends])) if col_ends is not None else ""
                planned_payment_date = excel_date_to_string(get_cell_value(row_cells[col_planned_payment])) if col_planned_payment is not None and get_cell_value(row_cells[col_planned_payment]) else ""
                
                paid_amount = clean_amount(get_cell_value(row_cells[col_paid_amount])) if col_paid_amount is not None else "0"
                beneficiary_bank_name = str(get_cell_value(row_cells[col_bank])).strip() if col_bank is not None and get_cell_value(row_cells[col_bank]) else ""
                
                # Create row
                csv_row = [
                    trip_number,
                    customer_name,
                    trip_destination,
                    reason_for_trip,
                    trip_begins_on,
                    trip_ends_on,
                    planned_payment_date,
                    paid_amount,
                    beneficiary_bank_name
                ]
                
                all_rows.append(csv_row)
                print(f"  Row {row_idx}: {trip_number} - {customer_name}")
                
            except Exception as e:
                print(f"  Error processing row {row_idx}: {e}")
                continue
    
    # Write to CSV
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(all_rows)
    
    print(f"\nSuccessfully converted {len(all_rows)} rows to {output_csv}")
    return len(all_rows)

if __name__ == "__main__":
    # You can switch between real file and sample file
    excel_file = r"D:\Bu Intan\Lampiran Ams - Tgl Bayar 17112025.XLSX"
    # excel_file = r"D:\Bu Intan\Bu Intan\data\sppd\sample_sppd_with_payment_date.xlsx"
    output_file = r"D:\Bu Intan\Bu Intan\data\sppd\sppd_lampiran_ams.csv"
    
    convert_excel_to_sppd_csv(excel_file, output_file)
